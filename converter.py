"""
PDF → Excel 변환기
PDF의 2D 레이아웃(x/y 좌표)을 Excel 그리드에 매핑합니다.
"""

import re
import io
import pdfplumber
import pytesseract
import fitz  # pymupdf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

_ILLEGAL_CHARS = re.compile(r'[\x01-\x08\x0b\x0c\x0e-\x1f]')

# 페이지 전체 너비를 나눌 Excel 컬럼 수
NUM_COLS = 20


# ── 유틸리티 ─────────────────────────────────────────────

def safe_value(v):
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v)
    # \x00이 단어 문자 사이에 있을 때만 하이픈으로, 나머지는 제거
    s = re.sub(r'(?<=\w)\x00(?=\w)', '-', s)
    s = s.replace('\x00', '')
    cleaned = _ILLEGAL_CHARS.sub('', s).strip()
    return cleaned if cleaned else None


def is_scanned_page(page) -> bool:
    return len((page.extract_text() or "").strip()) < 20


def ocr_page(image, lang="kor+eng") -> str:
    return pytesseract.image_to_string(image, lang=lang)


# ── PDF 좌표 → Excel 컬럼 매핑 ──────────────────────────

def x_to_col(x, page_width, num_cols=NUM_COLS):
    """PDF x 좌표를 Excel 컬럼 번호(1-indexed)로 변환"""
    col = int(x / page_width * num_cols) + 1
    return max(1, min(num_cols, col))


def bbox_to_cols(x0, x1, page_width, num_cols=NUM_COLS):
    col_start = x_to_col(x0, page_width, num_cols)
    col_end   = x_to_col(x1, page_width, num_cols)
    return col_start, max(col_start, col_end)


# ── 폰트 크기 정규화 ─────────────────────────────────────

def normalize_font_size(pt):
    """PDF 포인트 크기를 Excel에서 보기 좋은 크기로 정규화"""
    if not pt or pt <= 0:
        return 9
    if pt <= 8:
        return 9
    if pt <= 12:
        return round(pt)
    return min(round(pt * 0.75), 20)


# ── 셀 병합 안전 함수 ────────────────────────────────────

def safe_merge(ws, start_row, start_column, end_row, end_column):
    """기존 병합과 충돌하지 않을 때만 셀 병합."""
    if start_row == end_row and start_column == end_column:
        return
    try:
        ws.merge_cells(
            start_row=start_row, start_column=start_column,
            end_row=end_row,     end_column=end_column,
        )
    except Exception:
        pass


# ── 텍스트 정렬 추론 ─────────────────────────────────────

def infer_alignment(x0, x1, page_width):
    """bbox x 좌표로 텍스트 정렬 추론 (left/center/right)"""
    center_x = (x0 + x1) / 2
    if x0 < page_width * 0.08:
        return "left"
    if center_x > page_width * 0.62:
        return "right"
    if center_x > page_width * 0.32:
        return "center"
    return "left"


# ── 이미지 추출 ──────────────────────────────────────────

def crop_page_region(page_img, x0, top, x1, bottom, page_width, page_height):
    """PDF 좌표 영역을 페이지 렌더 이미지에서 크롭"""
    iw, ih = page_img.size
    sx, sy = iw / page_width, ih / page_height
    px0, py0 = int(x0 * sx), int(top * sy)
    px1, py1 = int(x1 * sx), int(bottom * sy)
    px0, px1 = max(0, px0), min(iw, px1)
    py0, py1 = max(0, py0), min(ih, py1)
    if px1 <= px0 or py1 <= py0:
        return None
    return page_img.crop((px0, py0, px1, py1))


# ── 병합 셀 감지 ─────────────────────────────────────────

def detect_spans(table):
    """
    수직 병합 먼저, 수평 병합 나중 처리.
    반환: {(row, col): [rowspan, colspan]}, merged_set
    """
    if not table:
        return {}, set()
    rows = len(table)
    cols = max((len(r) for r in table if r), default=0)
    if not cols:
        return {}, set()

    grid = [list(r) + [None] * (cols - len(r)) if r else [None] * cols for r in table]
    spans  = {}
    merged = set()

    # 비어있지 않은 모든 셀 초기화
    for i in range(rows):
        for j in range(cols):
            if grid[i][j] is not None:
                spans[(i, j)] = [1, 1]

    # 수직 병합 (같은 열, 아래쪽 None 흡수)
    for j in range(cols):
        for i in range(rows):
            if (i, j) not in spans:
                continue
            rs = 1
            while i + rs < rows and grid[i + rs][j] is None and (i + rs, j) not in merged:
                rs += 1
            if rs > 1:
                spans[(i, j)][0] = rs
                for k in range(1, rs):
                    merged.add((i + k, j))

    # 수평 병합 (같은 행, 오른쪽 None 흡수, 수직 병합된 셀 제외)
    for i in range(rows):
        for j in range(cols):
            if (i, j) not in spans:
                continue
            cs = 1
            while j + cs < cols and grid[i][j + cs] is None and (i, j + cs) not in merged:
                cs += 1
            if cs > 1:
                spans[(i, j)][1] = cs
                for k in range(1, cs):
                    merged.add((i, j + k))

    return spans, merged


# ── A4 페이지 테두리 ─────────────────────────────────────

def draw_page_border(ws, start_row, end_row, start_col, end_col):
    """A4 페이지 영역 외곽에 굵은 테두리를 그린다."""
    from openpyxl.cell.cell import MergedCell
    thick = Side(style='medium', color='000000')

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            is_top    = row == start_row
            is_bottom = row == end_row
            is_left   = col == start_col
            is_right  = col == end_col
            if not (is_top or is_bottom or is_left or is_right):
                continue
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            b = cell.border
            cell.border = Border(
                left   = thick if is_left   else b.left,
                right  = thick if is_right  else b.right,
                top    = thick if is_top    else b.top,
                bottom = thick if is_bottom else b.bottom,
            )


# ── 테이블 쓰기 ──────────────────────────────────────────

def write_table(ws, table, start_row, col_start=1, col_end=NUM_COLS, row_heights=None, col_boundaries=None):
    """테이블을 워크시트에 씁니다 (병합 셀 포함)."""
    if not table:
        return start_row

    rows = len(table)
    cols = max((len(r) for r in table if r), default=0)
    spans, merged = detect_spans(table)
    grid = [list(r) + [None] * (cols - len(r)) if r else [None] * cols for r in table]

    thin   = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_excel_cols = col_end - col_start + 1

    # PDF 컬럼 경계 좌표가 있으면 비례 매핑, 없으면 균등 분배
    if col_boundaries and len(col_boundaries) >= cols + 1:
        tbl_x0 = col_boundaries[0]
        tbl_width = col_boundaries[-1] - tbl_x0
        def tbl_col_to_excel(j):
            if j >= len(col_boundaries) - 1:
                return col_end + 1
            rel = (col_boundaries[j] - tbl_x0) / tbl_width
            return col_start + int(rel * total_excel_cols)
    else:
        def tbl_col_to_excel(j):
            return col_start + int(j / cols * total_excel_cols)

    for i, row in enumerate(grid):
        excel_row = start_row + i

        for j in range(cols):
            if (i, j) in merged:
                continue

            ec = tbl_col_to_excel(j)
            val = safe_value(row[j])
            rs, cs = spans.get((i, j), [1, 1])

            ec_end_row = excel_row + rs - 1
            if j + cs >= cols:
                ec_end_col = col_end
            else:
                ec_end_col = tbl_col_to_excel(j + cs) - 1
            ec_end_col = max(ec, min(col_end, ec_end_col))

            if ec_end_row > excel_row or ec_end_col > ec:
                safe_merge(ws, excel_row, ec, ec_end_row, ec_end_col)

            cell = ws.cell(row=excel_row, column=ec)
            cell.value = val
            cell.border = border
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # PDF 행 높이 반영, 셀 내 줄 수 기반 최소 높이 보장
        h = row_heights[i] if row_heights and i < len(row_heights) else 16
        max_lines = max((str(row[j] or '').count('\n') + 1 for j in range(cols)), default=1)
        min_h = max_lines * 14
        ws.row_dimensions[excel_row].height = max(min_h, h)

    return start_row + rows


# ── 레이아웃 요소 수집 ───────────────────────────────────

def collect_elements(page, page_img=None, lang="kor+eng"):
    """페이지에서 모든 요소(이미지, 텍스트, 테이블)를 bbox와 함께 수집."""
    elements = []
    table_bboxes = []

    # 테이블
    for tobj in page.find_tables():
        data = tobj.extract()
        if data:
            # 각 행의 PDF 높이 계산 (rows 속성 활용)
            row_heights = []
            try:
                for trow in tobj.rows:
                    h = trow.bbox[3] - trow.bbox[1]
                    row_heights.append(max(10, h))
            except Exception:
                row_heights = None
            col_boundaries = sorted(set(c[0] for c in tobj.cells)) + [tobj.bbox[2]]
            elements.append({'type': 'table', 'bbox': tobj.bbox, 'data': data,
                             'row_heights': row_heights, 'col_boundaries': col_boundaries})
            table_bboxes.append(tobj.bbox)

    # 이미지 (최소 크기 이상만)
    for img in page.images:
        w = img['x1'] - img['x0']
        h = img['bottom'] - img['top']
        if w < 20 or h < 20:
            continue
        pil_crop = None
        if page_img:
            pil_crop = crop_page_region(
                page_img, img['x0'], img['top'], img['x1'], img['bottom'],
                page.width, page.height
            )
        if pil_crop:
            elements.append({
                'type': 'image',
                'bbox': (img['x0'], img['top'], img['x1'], img['bottom']),
                'data': pil_crop,
            })

    # 텍스트 (테이블 외부, 한 줄씩 독립 요소로)
    words = page.extract_words(extra_attrs=['size'])
    non_tbl = [
        w for w in words
        if not any(
            b[0] <= w['x0'] and w['x1'] <= b[2] and b[1] <= w['top'] and w['bottom'] <= b[3]
            for b in table_bboxes
        )
    ]

    if non_tbl:
        lines, cur_line, cur_y = [], [], None
        for word in sorted(non_tbl, key=lambda w: (round(w['top']), w['x0'])):
            y = round(word['top'])
            if cur_y is None or abs(y - cur_y) > 4:
                if cur_line:
                    lines.append(cur_line)
                cur_line, cur_y = [word], y
            else:
                cur_line.append(word)
        if cur_line:
            lines.append(cur_line)

        gap_threshold = page.width / NUM_COLS * 1.5  # 약 1.5 컬럼 너비 이상 = 별도 요소
        for line in lines:
            # x 간격이 큰 곳에서 분리
            sorted_words = sorted(line, key=lambda w: w['x0'])
            groups, cur_group = [], [sorted_words[0]]
            for word in sorted_words[1:]:
                if word['x0'] - cur_group[-1]['x1'] > gap_threshold:
                    groups.append(cur_group)
                    cur_group = [word]
                else:
                    cur_group.append(word)
            groups.append(cur_group)

            for group in groups:
                x0  = min(w['x0']     for w in group)
                x1  = max(w['x1']     for w in group)
                top = min(w['top']    for w in group)
                bot = max(w['bottom'] for w in group)
                txt = ' '.join(w['text'] for w in group)
                sizes = [w['size'] for w in group if w.get('size')]
                font_size = normalize_font_size(max(sizes) if sizes else 0)
                alignment = infer_alignment(x0, x1, page.width)
                elements.append({
                    'type': 'text', 'bbox': (x0, top, x1, bot),
                    'data': txt, 'font_size': font_size, 'alignment': alignment,
                })

    # y 오름차순 → x 오름차순 정렬
    elements.sort(key=lambda e: (e['bbox'][1], e['bbox'][0]))
    return elements


def group_by_y_bands(elements, tolerance=10):
    """y 범위가 겹치는 요소들을 같은 밴드로 묶음 (테이블 제외)."""
    if not elements:
        return []
    bands, cur_band = [], [elements[0]]
    cur_bottom = elements[0]['bbox'][3]

    for elem in elements[1:]:
        y_top = elem['bbox'][1]
        if y_top <= cur_bottom + tolerance:
            cur_band.append(elem)
            cur_bottom = max(cur_bottom, elem['bbox'][3])
        else:
            bands.append(cur_band)
            cur_band, cur_bottom = [elem], elem['bbox'][3]
    bands.append(cur_band)
    return bands


def build_render_plan(elements):
    """
    테이블은 개별 항목으로, 텍스트/이미지는 y-밴드로 묶어
    y 순서대로 렌더링 계획을 반환합니다.
    """
    tables = [e for e in elements if e['type'] == 'table']
    others = [e for e in elements if e['type'] != 'table']

    items = []
    for t in tables:
        items.append({'kind': 'table', 'y': t['bbox'][1], 'elem': t})

    for band in (group_by_y_bands(others, tolerance=10) if others else []):
        y = min(e['bbox'][1] for e in band)
        items.append({'kind': 'band', 'y': y, 'band': band})

    items.sort(key=lambda x: x['y'])
    return items


# ── 메인 변환 ────────────────────────────────────────────

def convert_pdf_to_excel(pdf_path: str, lang: str = "kor+eng") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "변환 결과"

    current_row = 1

    with pdfplumber.open(pdf_path) as pdf:
        # 첫 페이지 너비 기준으로 컬럼 너비 비례 설정
        # PDF pt → Excel 문자 단위: A4(595pt) 기준 전체 너비 ≈ 85 chars
        first_page_width = pdf.pages[0].width if pdf.pages else 595
        total_excel_width = first_page_width * 0.185   # pt → Excel char 단위
        col_width = round(total_excel_width / NUM_COLS, 1)
        for c in range(1, NUM_COLS + 1):
            ws.column_dimensions[get_column_letter(c)].width = col_width
        # 스캔 페이지 감지 및 OCR 이미지 준비
        scanned = [i for i, p in enumerate(pdf.pages) if is_scanned_page(p)]
        page_imgs = {}

        # 이미지 렌더링 (스캔 OCR 또는 이미지 추출 목적)
        render_pages = set(scanned)
        for i, p in enumerate(pdf.pages):
            if p.images:
                render_pages.add(i)

        if render_pages:
            fitz_doc = fitz.open(pdf_path)
            for pi in render_pages:
                fitz_page = fitz_doc[pi]
                mat = fitz.Matrix(180 / 72, 180 / 72)
                pix = fitz_page.get_pixmap(matrix=mat)
                page_imgs[pi] = PILImage.frombytes("RGB", [pix.width, pix.height], pix.samples)
            fitz_doc.close()

        for page_num, page in enumerate(pdf.pages, start=1):
            pi = page_num - 1
            page_img = page_imgs.get(pi)
            page_start_row = current_row
            # A4 기준 y좌표 → Excel 행 매핑: 14pt 행 높이 기준
            ROW_HEIGHT_PT = 14
            rows_per_page = max(60, int(page.height / ROW_HEIGHT_PT))

            def y_to_row(y):
                return page_start_row + max(0, int(y / page.height * rows_per_page))

            if is_scanned_page(page) and page_img:
                text = ocr_page(page_img, lang)
                for line in text.split('\n'):
                    line = line.strip()
                    if line:
                        ws.cell(row=current_row, column=1, value=safe_value(line))
                        ws.row_dimensions[current_row].height = 15
                        current_row += 1
            else:
                elements = collect_elements(page, page_img, lang)
                plan = build_render_plan(elements)

                for item in plan:
                    # y좌표 기반으로 목표 행 결정 (뒤로 가지 않음)
                    target_row = y_to_row(item['y'])
                    current_row = max(current_row, target_row)

                    if item['kind'] == 'table':
                        elem = item['elem']
                        col_s, col_e = bbox_to_cols(elem['bbox'][0], elem['bbox'][2], page.width, NUM_COLS)
                        end_row = write_table(ws, elem['data'], current_row, col_s, col_e,
                                              row_heights=elem.get('row_heights'),
                                              col_boundaries=elem.get('col_boundaries'))
                        current_row = end_row + 1

                    elif item['kind'] == 'band':
                        band = item['band']
                        text_elems = [e for e in band if e['type'] == 'text']
                        img_elems  = [e for e in band if e['type'] == 'image']

                        # 텍스트 줄별 y → 서브행 (같은 band 내에서만)
                        unique_tops = sorted(set(round(e['bbox'][1]) for e in text_elems))
                        top_to_subrow = {t: i for i, t in enumerate(unique_tops)}
                        text_rows = len(unique_tops) if unique_tops else 1

                        subrow_height = {}
                        for e in text_elems:
                            si = top_to_subrow.get(round(e['bbox'][1]), 0)
                            h = max(10, e['bbox'][3] - e['bbox'][1])
                            subrow_height[si] = max(subrow_height.get(si, 0), h)

                        img_rows = 1
                        img_total_h = max((e['bbox'][3] - e['bbox'][1] for e in img_elems), default=0)
                        if img_total_h > 0:
                            img_rows = max(1, int(img_total_h / ROW_HEIGHT_PT))

                        band_rows = max(text_rows, img_rows)

                        for elem in text_elems:
                            col_s, col_e = bbox_to_cols(elem['bbox'][0], elem['bbox'][2], page.width, NUM_COLS)
                            # 우측 정렬 텍스트: 빈 셀 확인 후 왼쪽 확장
                            if elem.get('alignment') == 'right' and col_e >= NUM_COLS - 3:
                                from openpyxl.cell.cell import MergedCell as MC
                                subrow_tmp = top_to_subrow.get(round(elem['bbox'][1]), 0)
                                r_tmp = current_row + subrow_tmp
                                txt_len = len(str(elem['data']))
                                needed = max(4, txt_len // 4 + 2)
                                new_col_s = max(1, col_e - needed + 1)
                                candidate = ws.cell(row=r_tmp, column=new_col_s)
                                if not isinstance(candidate, MC) and candidate.value is None:
                                    col_s = new_col_s
                                col_e = NUM_COLS
                            subrow = top_to_subrow.get(round(elem['bbox'][1]), 0)
                            r = current_row + subrow
                            if col_e > col_s:
                                safe_merge(ws, r, col_s, r, col_e)
                            raw_cell = ws.cell(row=r, column=col_s)
                            from openpyxl.cell.cell import MergedCell
                            if isinstance(raw_cell, MergedCell):
                                continue
                            raw_cell.value = safe_value(elem['data'])
                            raw_cell.font = Font(size=elem.get('font_size', 9))
                            raw_cell.alignment = Alignment(
                                horizontal=elem.get('alignment', 'left'),
                                vertical="center", wrap_text=True
                            )

                        for elem in img_elems:
                            col_s, col_e = bbox_to_cols(elem['bbox'][0], elem['bbox'][2], page.width, NUM_COLS)
                            img_buf = io.BytesIO()
                            elem['data'].save(img_buf, format='PNG')
                            img_buf.seek(0)
                            xl_img = XLImage(img_buf)
                            w_pt = elem['bbox'][2] - elem['bbox'][0]
                            h_pt = elem['bbox'][3] - elem['bbox'][1]
                            xl_img.width  = int(w_pt * 1.8)
                            xl_img.height = int(h_pt * 1.8)
                            ws.add_image(xl_img, f"{get_column_letter(col_s)}{current_row}")

                        for si in range(band_rows):
                            row_h = subrow_height.get(si, ROW_HEIGHT_PT)
                            ws.row_dimensions[current_row + si].height = max(10, row_h)

                        current_row += band_rows

            # A4 페이지 영역 테두리
            page_end_row = page_start_row + rows_per_page - 1
            draw_page_border(ws, page_start_row, page_end_row, 1, NUM_COLS)

            # 다음 페이지는 현재 페이지 끝 행 이후부터
            current_row = max(current_row, page_start_row + rows_per_page) + 2

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
